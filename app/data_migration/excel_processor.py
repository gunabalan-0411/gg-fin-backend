"""
Reuses the Excel processing logic from the original notebook.
Reads Excel files → Pandas DataFrames with validated columns.
"""
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


class ExcelCollectionProcessor:
    def __init__(self, file_path: str, sheet_name: str, file_type: str):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.file_type = file_type.lower()

    def classify_cell(self, cell):
        payment_mode = "CASH"
        payment_status = "PAID"
        color = cell.fill.start_color
        if color.type == "theme":
            if color.theme in (4, 8):
                payment_mode = "ONLINE"
                payment_status = "PENDING"
            elif color.theme == 9:
                payment_mode = "ONLINE"
                payment_status = "PAID"
        return payment_mode, payment_status

    def transform_transactions(self) -> pd.DataFrame:
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb[self.sheet_name]

        if self.file_type == "edi":
            start_row, date_col, start_col = 2, 2, 3
        elif self.file_type == "iop":
            start_row, date_col, start_col = 3, 2, 4
        else:
            raise ValueError("file_type must be 'edi' or 'iop'")

        records = []
        for r in tqdm(range(start_row, ws.max_row + 1)):
            collection_date = ws.cell(r, date_col).value
            for c in range(start_col, ws.max_column + 1):
                customer_id = ws.cell(1, c).value
                cell = ws.cell(r, c)
                amount = cell.value
                payment_mode, payment_status = self.classify_cell(cell)
                records.append({
                    "customer_id": customer_id,
                    "collection_date": collection_date,
                    "amount": amount,
                    "payment_mode": payment_mode,
                    "payment_status": payment_status,
                })

        df = pd.DataFrame(records)
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
        df = df[df["amount"].fillna(0) > 0].reset_index(drop=True)
        df.insert(0, "transaction_id", range(1, len(df) + 1))
        df["collection_date"] = pd.to_datetime(df["collection_date"]).dt.date
        return df

    def transform_edi_customer_details(self) -> pd.DataFrame:
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, dtype={"Phone": "string"})
        df = df[["ID", "Month", "Date", "Group", "Name", "Address", "aadhaar",
                  "Phone", "Loan", "Amount to customer", "Interest amount", "Balance",
                  "Note", "Unnamed: 15", "Unnamed: 16", "Unnamed: 17"]]
        note_cols = ["Note", "Unnamed: 15", "Unnamed: 16", "Unnamed: 17"]
        df["Notes"] = df[note_cols].fillna("").astype(str).agg(" ".join, axis=1)
        df = df.drop(columns=note_cols)
        df.columns = [
            "customer_id", "month", "loan_start_date", "customer_segment_id",
            "customer_name", "customer_address", "proof_aadhaar", "contact_number",
            "loan_amount", "disbursed_amount", "interest", "outstanding_balance", "remarks",
        ]
        df["loan_start_date"] = pd.to_datetime(df["loan_start_date"], errors="coerce").dt.date
        return df

    def transform_iop_customer_details(self) -> pd.DataFrame:
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, dtype={"Phone": "string"})
        note_cols = [c for c in ["Notes", "Unnamed: 14", "Unnamed: 15",
                                  "Unnamed: 16", "Unnamed: 17", "Unnamed: 18"] if c in df.columns]
        df["Notes"] = df[note_cols].apply(
            lambda row: " ".join(
                str(v).strip() for v in row.values
                if pd.notna(v) and str(v).strip().lower() not in ["nan", "nat"]
            ),
            axis=1,
        )
        df = df.drop(columns=[c for c in note_cols if c != "Notes"])
        df.columns = [
            "customer_id", "month", "loan_start_date", "customer_segment_id",
            "customer_name", "customer_address", "proof_aadhaar", "contact_number",
            "interest_payment_frequency", "loan_amount", "disbursed_amount",
            "interest", "_loan_closure_legacy", "remarks",
        ]
        df = df.drop(columns=["_loan_closure_legacy"])
        df["loan_start_date"] = pd.to_datetime(df["loan_start_date"], errors="coerce").dt.date
        return df

    def transform_mapping_tbls(self, columns: list[str]) -> pd.DataFrame:
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        df.columns = columns
        return df

    def transform_expense_tbl(self) -> pd.DataFrame:
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        note_cols = [c for c in df.columns if c.startswith("Note") or c.startswith("Unnamed")]
        df["Notes"] = df[note_cols].apply(
            lambda row: ", ".join(
                str(v).strip() for v in row.values
                if pd.notna(v) and str(v).strip().lower() not in ["nan", "nat"]
            ),
            axis=1,
        )
        df = df.drop(columns=[c for c in note_cols if c in df.columns])
        df.columns = ["amount", "date", "notes"]
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
        df = df.dropna(subset=["amount", "date"])
        return df
